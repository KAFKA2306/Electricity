#!/usr/bin/env python3
"""Publish keyless EIA electricity evidence and derived views."""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

ROOT = Path(__file__).resolve().parents[1]
EBA_URL = "https://www.eia.gov/opendata/bulk/EBA.zip"
EIA860M_INDEX = "https://www.eia.gov/electricity/data/eia860m/"
TOTAL_SERIES = {"demand": "EBA.US48-ALL.D.H", "generation": "EBA.US48-ALL.NG.H"}
FUEL_RE = re.compile(r"^EBA\.US48-ALL\.NG\.([A-Z0-9]+)\.H$")
INTERCHANGE_ID = "EBA.US48-ALL.TI.H"
SHEETS = ("Operating", "Planned", "Retired")
UA = "energy-supply/1.0 github.com/KAFKA2306/oil"


def get(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as response:
        return response.read()


def sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def dump(value: object) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode()


class Links(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        href = dict(attrs).get("href")
        if tag.lower() == "a" and href:
            self.hrefs.append(href)


def latest_eia860m_url(html: bytes) -> str:
    parser = Links()
    parser.feed(html.decode(errors="replace"))
    urls = [
        urljoin(EIA860M_INDEX, href)
        for href in parser.hrefs
        if re.search(r"/eia860m/xls/[a-z]+_generator20\d{2}\.xlsx$", href, re.I)
    ]
    if not urls:
        raise ValueError("EIA-860M generator workbook not found")
    return urls[0]


def read_us48_bulk(zip_bytes: bytes) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, dict[str, Any]],
    dict[str, dict[str, Any]],
]:
    totals: dict[str, dict[str, Any]] = {}
    fuels: dict[str, dict[str, Any]] = {}
    interchange: dict[str, dict[str, Any]] = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        names = [name for name in archive.namelist() if name.lower().endswith(".txt")]
        if len(names) != 1:
            raise ValueError(f"expected one EBA txt file, got {names}")
        with archive.open(names[0]) as source:
            for line in source:
                if b"EBA.US48-ALL." not in line:
                    continue
                obj = json.loads(line)
                sid = str(obj.get("series_id") or "")
                if sid in TOTAL_SERIES.values():
                    totals[sid] = obj
                match = FUEL_RE.match(sid)
                if match:
                    fuels[match.group(1)] = obj
                if sid == INTERCHANGE_ID or (
                    sid.startswith("EBA.US48-ALL.")
                    and "interchange" in str(obj.get("name") or "").lower()
                    and obj.get("f") == "H"
                ):
                    interchange[sid] = obj
    missing = set(TOTAL_SERIES.values()) - totals.keys()
    if missing:
        raise ValueError(f"EBA bulk missing totals {sorted(missing)}")
    if len(fuels) < 5:
        raise ValueError(
            f"EBA bulk exposed only {len(fuels)} US48 fuel-generation series: {sorted(fuels)}"
        )
    if not interchange:
        raise ValueError("EBA bulk exposed no US48 hourly interchange series")
    return totals, fuels, interchange


def parse_hour(value: str) -> datetime:
    return datetime.strptime(value.removesuffix("Z"), "%Y%m%dT%H").replace(tzinfo=timezone.utc)


def daily_actual(series: dict[str, Any], days: int = 120) -> list[dict[str, Any]]:
    if series.get("f") != "H":
        raise ValueError(f"{series.get('series_id')}: expected hourly data")
    points: list[tuple[datetime, float]] = []
    for period, value in series.get("data") or []:
        try:
            points.append((parse_hour(str(period)), float(value)))
        except (TypeError, ValueError):
            continue
    if not points:
        raise ValueError(f"{series.get('series_id')}: no numeric hourly observations")
    latest = max(dt.date() for dt, _ in points)
    cutoff = latest - timedelta(days=days - 1)
    grouped: dict[str, list[float]] = defaultdict(list)
    for dt, value in points:
        if dt.date() >= cutoff:
            grouped[dt.date().isoformat()].append(value)
    rows = [
        {"period": day, "value": round(sum(values), 3), "hour_count": len(values)}
        for day, values in sorted(grouped.items())
        if len(values) >= 20
    ]
    if len(rows) < 90:
        raise ValueError(
            f"{series.get('series_id')}: expected >=90 complete days, got {len(rows)}"
        )
    return rows


def weekly(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        day = date.fromisoformat(row["period"])
        monday = (day - timedelta(days=day.weekday())).isoformat()
        groups[monday].append(float(row["value"]))
    return [
        {"week_start": start, "value": round(sum(values), 3), "day_count": len(values)}
        for start, values in sorted(groups.items())
    ]


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def worksheet_records(ws: Any) -> list[dict[str, Any]]:
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [str(value or "").strip() for value in row]
        normalized = [norm(value) for value in row]
        if "generator id" in normalized and (
            "plant id" in normalized or "plant code" in normalized
        ):
            return [
                {
                    headers[i]: value
                    for i, value in enumerate(record)
                    if i < len(headers) and headers[i]
                }
                for record in ws.iter_rows(min_row=row_number + 1, values_only=True)
                if any(value not in (None, "") for value in record)
            ]
        if row_number >= 15:
            break
    raise ValueError(f"{ws.title}: generator header not found")


def column(record: dict[str, Any], *names: str) -> str:
    keys = {norm(key): key for key in record}
    for name in names:
        if norm(name) in keys:
            return keys[norm(name)]
    raise ValueError(f"missing column {names}; sample={list(record)[:20]}")


def number(value: object) -> float | None:
    try:
        return None if value in (None, "", "NA", "N/A") else float(value)
    except (TypeError, ValueError):
        return None


def capacity_summary(xlsx: bytes) -> dict[str, Any]:
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    if any(name not in book.sheetnames for name in SHEETS):
        raise ValueError(f"EIA-860M sheets changed: {book.sheetnames}")
    result: dict[str, Any] = {}
    for sheet_name in SHEETS:
        records = worksheet_records(book[sheet_name])
        sample = records[0]
        summer = column(sample, "Net Summer Capacity (MW)")
        nameplate = column(sample, "Nameplate Capacity (MW)")
        technology = column(sample, "Technology")
        by_technology: dict[str, float] = defaultdict(float)
        summer_total = nameplate_total = 0.0
        for record in records:
            s = number(record.get(summer))
            n = number(record.get(nameplate))
            if s is not None:
                summer_total += s
                by_technology[str(record.get(technology) or "Unknown")] += s
            if n is not None:
                nameplate_total += n
        result[sheet_name.lower()] = {
            "generator_count": len(records),
            "nameplate_capacity_mw": round(nameplate_total, 3),
            "net_summer_capacity_mw": round(summer_total, 3),
            "by_technology_net_summer_mw": dict(sorted(by_technology.items())),
        }
    return result


def series_view(name: str, source: dict[str, Any], days: int) -> dict[str, Any]:
    daily = daily_actual(source, days)
    return {
        "schema_version": 1,
        "dataset": name,
        "series_id": source["series_id"],
        "frequency": "daily",
        "kind": "actual",
        "unit": source.get("units") or "megawatthours",
        "day_boundary": "UTC",
        "geography": {
            "level": "region",
            "id": "US48",
            "name": "United States Lower 48",
        },
        "first_period": daily[0]["period"],
        "last_period": daily[-1]["period"],
        "period_count": len(daily),
        "daily": daily,
        "weekly": weekly(daily),
        "source_url": EBA_URL,
    }


def build(days: int) -> dict[str, Any]:
    eba = get(EBA_URL)
    totals, fuels, interchanges = read_us48_bulk(eba)
    capacity_url = latest_eia860m_url(get(EIA860M_INDEX))
    xlsx = get(capacity_url)

    views = {
        name: series_view(name, totals[series_id], days)
        for name, series_id in TOTAL_SERIES.items()
    }
    interchange_source = interchanges.get(INTERCHANGE_ID) or next(iter(interchanges.values()))
    views["interchange"] = series_view("interchange", interchange_source, days)

    fuel_views: dict[str, Any] = {}
    excluded: dict[str, str] = {}
    for code, source in sorted(fuels.items()):
        try:
            fuel_views[code] = series_view(f"generation-{code.lower()}", source, days)
        except ValueError as exc:
            excluded[code] = str(exc)
    if len(fuel_views) < 5:
        raise ValueError(
            f"fewer than five 90-day fuel series survived: {sorted(fuel_views)}; excluded={excluded}"
        )

    common = set.intersection(
        *(set(row["period"] for row in item["daily"]) for item in fuel_views.values())
    )
    if not common:
        raise ValueError("fuel-generation series have no common daily period")
    latest = max(common)
    latest_values = {
        code: next(row["value"] for row in item["daily"] if row["period"] == latest)
        for code, item in fuel_views.items()
    }
    total = sum(latest_values.values())
    generation_mix = {
        "schema_version": 1,
        "dataset": "generation-mix",
        "frequency": "daily",
        "kind": "actual",
        "unit": "megawatthours",
        "day_boundary": "UTC",
        "geography": {
            "level": "region",
            "id": "US48",
            "name": "United States Lower 48",
        },
        "fuel_count": len(fuel_views),
        "fuels": fuel_views,
        "excluded_series": excluded,
        "latest": {
            "period": latest,
            "total_mwh": round(total, 3),
            "by_fuel_mwh": latest_values,
            "share": {
                code: round(value / total, 8) for code, value in latest_values.items()
            }
            if total
            else {},
        },
        "source_url": EBA_URL,
    }

    return {
        "schema_version": 4,
        "publisher": "U.S. Energy Information Administration",
        "retrieved_at": datetime.now(timezone.utc).isoformat(),
        "sources": {
            "eba_bulk": {"url": EBA_URL, "sha256": sha(eba)},
            "eia860m": {
                "index_url": EIA860M_INDEX,
                "url": capacity_url,
                "sha256": sha(xlsx),
                "sheets": list(SHEETS),
            },
        },
        "views": views,
        "generation_mix": generation_mix,
        "capacity": capacity_summary(xlsx),
    }


def evidence_name(payload: dict[str, Any]) -> str:
    stable = dict(payload)
    stable.pop("retrieved_at", None)
    return f"eia-electricity-{sha(dump(stable))[:16]}.json"


def publish(
    evidence_dir: Path, api_dir: Path, payload: dict[str, Any]
) -> dict[str, Any]:
    evidence_dir.mkdir(parents=True, exist_ok=True)
    evidence = evidence_dir / evidence_name(payload)
    if not evidence.exists():
        evidence.write_bytes(dump(payload))
    stored = json.loads(evidence.read_text())
    evidence_sha = sha(evidence.read_bytes())
    evidence_path = (
        str(evidence.relative_to(ROOT)) if evidence.is_relative_to(ROOT) else str(evidence)
    )
    api_dir.mkdir(parents=True, exist_ok=True)

    published: dict[str, Any] = {}
    for name, view in stored["views"].items():
        published[name] = {
            **view,
            "source_evidence": evidence_path,
            "source_evidence_sha256": evidence_sha,
        }
    published["generation-mix"] = {
        **stored["generation_mix"],
        "source_evidence": evidence_path,
        "source_evidence_sha256": evidence_sha,
    }
    published["capacity"] = {
        "schema_version": 1,
        "dataset": "capacity",
        "frequency": "monthly",
        "kind": "reported-inventory",
        "unit": "MW",
        "geography": {
            "level": "generator",
            "coverage": "United States and Puerto Rico where reported",
        },
        **stored["capacity"],
        "source_url": stored["sources"]["eia860m"]["url"],
        "source_evidence": evidence_path,
        "source_evidence_sha256": evidence_sha,
    }
    published["capacity-additions"] = {
        "schema_version": 1,
        "dataset": "capacity-additions",
        "frequency": "monthly",
        "kind": "reported-inventory",
        "unit": "MW",
        "planned": stored["capacity"]["planned"],
        "retired": stored["capacity"]["retired"],
        "source_url": stored["sources"]["eia860m"]["url"],
        "source_evidence": evidence_path,
        "source_evidence_sha256": evidence_sha,
    }

    for name, view in published.items():
        (api_dir / f"{name}.json").write_bytes(dump(view))
    (api_dir / "index.json").write_bytes(
        dump(
            {
                "schema_version": 1,
                "publisher": stored["publisher"],
                "namespace": "electricity",
                "datasets": {
                    name: {
                        "path": f"{name}.json",
                        "frequency": view["frequency"],
                        "kind": view["kind"],
                        "source_url": view["source_url"],
                    }
                    for name, view in sorted(published.items())
                },
            }
        )
    )
    (api_dir.parent / "energy-supply.json").write_bytes(
        dump(
            {
                "schema_version": 1,
                "theme": "ARK Big Ideas 2026 / Distributed Energy",
                "domains": {
                    "petroleum": {"namespace": "petroleum", "path": "latest.json"},
                    "electricity": {
                        "namespace": "electricity",
                        "path": "electricity/index.json",
                    },
                },
                "rule": "Domains are linked by reference only; units and values are never merged.",
            }
        )
    )
    return published


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=120)
    parser.add_argument(
        "--evidence-dir", type=Path, default=ROOT / "data/electricity/official"
    )
    parser.add_argument("--api-dir", type=Path, default=ROOT / "api/v1/electricity")
    args = parser.parse_args()
    if args.days < 90:
        raise SystemExit("--days must be >=90")
    views = publish(args.evidence_dir, args.api_dir, build(args.days))
    print(
        json.dumps(
            {
                "demand_days": views["demand"]["period_count"],
                "generation_days": views["generation"]["period_count"],
                "interchange_days": views["interchange"]["period_count"],
                "fuel_count": views["generation-mix"]["fuel_count"],
                "operating_generators": views["capacity"]["operating"]["generator_count"],
                "planned_generators": views["capacity"]["planned"]["generator_count"],
                "retired_generators": views["capacity"]["retired"]["generator_count"],
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
